import os
import bcrypt
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import db, Admin, Grupo, Estudiante, Jurado, Criterio, Calificacion
from config import Config
from datetime import datetime, timezone
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.config.from_object(Config)
db.init_app(app)

login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor inicia sesión para acceder.'
login_manager.init_app(app)

@login_manager.user_loader
def load_user(user_id):
    if user_id.startswith('a_'):
        return db.session.get(Admin, int(user_id[2:]))
    elif user_id.startswith('j_'):
        return db.session.get(Jurado, int(user_id[2:]))
    return None

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not isinstance(current_user, Admin):
            flash('Acceso denegado. Se requieren permisos de administrador.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def init_db():
    with app.app_context():
        db.create_all()
        if Criterio.query.count() == 0:
            criterios = [
                Criterio(nombre='Investigaci\u00f3n y concepto', componente='Portafolio', porcentaje=10, orden=1),
                Criterio(nombre='Ilustraciones: Figurines', componente='Portafolio', porcentaje=10, orden=2),
                Criterio(nombre='Fichas T\u00e9cnicas', componente='Portafolio', porcentaje=10, orden=3),
                Criterio(nombre='Vocabulario t\u00e9cnico y seguridad', componente='Sustentaci\u00f3n', porcentaje=10, orden=4),
                Criterio(nombre='Presentaci\u00f3n y Est\u00e9tica', componente='Sustentaci\u00f3n', porcentaje=10, orden=5),
                Criterio(nombre='Presentaci\u00f3n Personal', componente='Sustentaci\u00f3n', porcentaje=5, orden=6),
                Criterio(nombre='Lookbook o V\u00eddeo', componente='Sustentaci\u00f3n', porcentaje=5, orden=7),
                Criterio(nombre='Precisi\u00f3n y claridad t\u00e9cnica en el patr\u00f3n', componente='Runway', porcentaje=10, orden=8),
                Criterio(nombre='Materializaci\u00f3n: intervenci\u00f3n textil y estructura', componente='Runway', porcentaje=10, orden=9),
                Criterio(nombre='Estilismo y Coherencia Visual "Total Look"', componente='Runway', porcentaje=10, orden=10),
                Criterio(nombre='Desempe\u00f1o en Backstage y Disciplina', componente='Runway', porcentaje=10, orden=11),
            ]
            db.session.add_all(criterios)
            db.session.commit()

def calcular_notas(estudiante_id, jurado_id):
    calificaciones = Calificacion.query.filter_by(
        estudiante_id=estudiante_id, jurado_id=jurado_id
    ).all()
    if not calificaciones:
        return None
    calif_dict = {c.criterio_id: c.nota for c in calificaciones}
    componentes = {'Portafolio': 0.30, 'Sustentaci\u00f3n': 0.30, 'Runway': 0.40}
    notas_componente = {}
    for comp, peso_global in componentes.items():
        criterios = Criterio.query.filter_by(componente=comp).order_by(Criterio.orden).all()
        suma_ponderada = 0
        suma_porcentajes = 0
        for c in criterios:
            if c.id in calif_dict:
                suma_ponderada += calif_dict[c.id] * c.porcentaje
                suma_porcentajes += c.porcentaje
        if suma_porcentajes > 0:
            notas_componente[comp] = round(suma_ponderada / suma_porcentajes, 2)
        else:
            notas_componente[comp] = 0
    nota_final = round(
        notas_componente.get('Portafolio', 0) * 0.30 +
        notas_componente.get('Sustentaci\u00f3n', 0) * 0.30 +
        notas_componente.get('Runway', 0) * 0.40,
        2
    )
    return {'notas_componente': notas_componente, 'nota_final': nota_final}

def calcular_promedio_jurados(estudiante_id):
    jurados = Jurado.query.all()
    notas_finales = []
    for j in jurados:
        resultado = calcular_notas(estudiante_id, j.id)
        if resultado:
            notas_finales.append({'jurado': j, 'resultado': resultado})
    if not notas_finales:
        return None
    prom_portafolio = round(sum(nf['resultado']['notas_componente'].get('Portafolio', 0) for nf in notas_finales) / len(notas_finales), 2)
    prom_sustentacion = round(sum(nf['resultado']['notas_componente'].get('Sustentaci\u00f3n', 0) for nf in notas_finales) / len(notas_finales), 2)
    prom_runway = round(sum(nf['resultado']['notas_componente'].get('Runway', 0) for nf in notas_finales) / len(notas_finales), 2)
    prom_final = round(
        prom_portafolio * 0.30 + prom_sustentacion * 0.30 + prom_runway * 0.40,
        2
    )
    return {
        'notas_finales': notas_finales,
        'promedio_portafolio': prom_portafolio,
        'promedio_sustentacion': prom_sustentacion,
        'promedio_runway': prom_runway,
        'promedio_final': prom_final
    }

def get_grupos_con_estudiantes():
    grupos = Grupo.query.order_by(Grupo.nombre).all()
    result = []
    for g in grupos:
        estudiantes = Estudiante.query.filter_by(grupo_id=g.id).order_by(Estudiante.nombre).all()
        result.append({'grupo': g, 'estudiantes': estudiantes})
    return result

# ─── Rutas p├║blicas ───

@app.route('/')
def index():
    grupos_con_estudiantes = get_grupos_con_estudiantes()
    jurados = Jurado.query.order_by(Jurado.nombre).all()
    total_estudiantes = sum(len(item['estudiantes']) for item in grupos_con_estudiantes)
    return render_template('index.html', grupos_con_estudiantes=grupos_con_estudiantes,
                           jurados=jurados, total_estudiantes=total_estudiantes)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        tipo = request.form.get('tipo', 'jurado')
        if tipo == 'admin':
            username = request.form.get('username', '')
            password = request.form.get('password', '')
            admin = Admin.query.filter_by(username=username).first()
            if admin and bcrypt.checkpw(password.encode('utf-8'), admin.password.encode('utf-8')):
                login_user(admin)
                flash(f'Bienvenido administrador {admin.nombre}', 'success')
                return redirect(url_for('admin_index'))
            else:
                flash('Credenciales de administrador incorrectas', 'error')
        else:
            jurado_id = request.form.get('jurado_id')
            password = request.form.get('password')
            jurado = db.session.get(Jurado, int(jurado_id)) if jurado_id and jurado_id.isdigit() else None
            if jurado and jurado.password:
                if bcrypt.checkpw(password.encode('utf-8'), jurado.password.encode('utf-8')):
                    login_user(jurado)
                    flash(f'Bienvenido/a {jurado.nombre}', 'success')
                    return redirect(url_for('index'))
                else:
                    flash('Contrase├▒a incorrecta', 'error')
            elif jurado and not jurado.password:
                login_user(jurado)
                flash(f'Bienvenido/a {jurado.nombre}', 'success')
                return redirect(url_for('index'))
            else:
                flash('Jurado no encontrado', 'error')
    jurados = Jurado.query.order_by(Jurado.nombre).all()
    return render_template('login.html', jurados=jurados)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Sesi├│n cerrada correctamente', 'success')
    return redirect(url_for('login'))

@app.route('/estudiantes')
def listar_estudiantes():
    grupos = Grupo.query.order_by(Grupo.nombre).all()
    grupos_con_estudiantes = get_grupos_con_estudiantes()
    return render_template('estudiantes.html', grupos_con_estudiantes=grupos_con_estudiantes)

@app.route('/calificar/<int:estudiante_id>', methods=['GET', 'POST'])
@login_required
def calificar_estudiante(estudiante_id):
    if not isinstance(current_user, Jurado):
        flash('Solo los jurados pueden calificar.', 'error')
        return redirect(url_for('index'))
    estudiante = db.session.get(Estudiante, estudiante_id)
    if not estudiante:
        flash('Estudiante no encontrado', 'error')
        return redirect(url_for('index'))
    criterios = Criterio.query.order_by(Criterio.orden).all()
    if request.method == 'POST':
        todas_validas = True
        for c in criterios:
            nota_key = f'nota_{c.id}'
            if nota_key in request.form and request.form[nota_key].strip() != '':
                try:
                    nota = float(request.form[nota_key])
                    if nota < 0 or nota > 5:
                        todas_validas = False
                        continue
                    existing = Calificacion.query.filter_by(
                        estudiante_id=estudiante.id, criterio_id=c.id, jurado_id=current_user.id
                    ).first()
                    if existing:
                        existing.nota = nota
                        existing.fecha = datetime.now(timezone.utc)
                    else:
                        cal = Calificacion(
                            estudiante_id=estudiante.id, criterio_id=c.id,
                            jurado_id=current_user.id, nota=nota
                        )
                        db.session.add(cal)
                except ValueError:
                    todas_validas = False
            else:
                todas_validas = False
        if todas_validas:
            db.session.commit()
            flash(f'Calificaciones guardadas para {estudiante.nombre}', 'success')
            return redirect(url_for('index'))
        db.session.rollback()
        flash('Error. Verifica que todas las notas est\u00e9n entre 0.0 y 5.0.', 'error')
    calificaciones_existentes = Calificacion.query.filter_by(
        estudiante_id=estudiante.id, jurado_id=current_user.id
    ).all()
    calif_dict = {c.criterio_id: c.nota for c in calificaciones_existentes}
    criterios_por_componente = {}
    for c in criterios:
        criterios_por_componente.setdefault(c.componente, []).append(c)
    componentes_orden = {'Portafolio': 0, 'Sustentaci\u00f3n': 1, 'Runway': 2}
    criterios_por_componente = dict(
        sorted(criterios_por_componente.items(), key=lambda x: componentes_orden.get(x[0], 99))
    )
    return render_template('calificar.html', estudiante=estudiante,
                           criterios_por_componente=criterios_por_componente,
                           calif_dict=calif_dict, total_criterios=len(criterios))

@app.route('/reporte')
def reporte():
    grupos = Grupo.query.order_by(Grupo.nombre).all()
    jurados = Jurado.query.order_by(Jurado.nombre).all()
    datos = []
    for g in grupos:
        estudiantes = Estudiante.query.filter_by(grupo_id=g.id).order_by(Estudiante.nombre).all()
        grupo_datos = []
        for e in estudiantes:
            resultado = calcular_promedio_jurados(e.id)
            if resultado:
                grupo_datos.append({'estudiante': e, 'resultado': resultado})
        if grupo_datos:
            datos.append({'grupo': g, 'estudiantes': grupo_datos})
    return render_template('reporte.html', datos=datos, jurados=jurados)

@app.route('/exportar_excel')
def exportar_excel():
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0033A0', end_color='0033A0', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    ws1 = wb.active
    ws1.title = 'Notas finales'
    headers1 = ['Estudiante', 'C\u00f3digo', 'Grupo', 'Nota Portafolio', 'Nota Sustentaci\u00f3n', 'Nota Runway', 'Nota Final']
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    estudiantes = Estudiante.query.order_by(Estudiante.nombre).all()
    row = 2
    for e in estudiantes:
        resultado = calcular_promedio_jurados(e.id)
        ws1.cell(row=row, column=1, value=e.nombre).border = thin_border
        ws1.cell(row=row, column=2, value=e.codigo).border = thin_border
        ws1.cell(row=row, column=3, value=e.grupo_obj.nombre if e.grupo_obj else '').border = thin_border
        if resultado:
            ws1.cell(row=row, column=4, value=resultado['promedio_portafolio']).border = thin_border
            ws1.cell(row=row, column=5, value=resultado['promedio_sustentacion']).border = thin_border
            ws1.cell(row=row, column=6, value=resultado['promedio_runway']).border = thin_border
            ws1.cell(row=row, column=7, value=resultado['promedio_final']).border = thin_border
        else:
            for col in range(4, 8):
                ws1.cell(row=row, column=col, value='Sin notas').border = thin_border
        row += 1
    for col in range(1, 8):
        ws1.column_dimensions[chr(64 + col)].width = 22
    ws2 = wb.create_sheet('Detalle por jurado')
    headers2 = ['Estudiante', 'Jurado', 'Criterio', 'Componente', 'Nota']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    calificaciones = Calificacion.query.order_by(Calificacion.estudiante_id, Calificacion.jurado_id, Calificacion.criterio_id).all()
    row = 2
    for cal in calificaciones:
        ws2.cell(row=row, column=1, value=cal.estudiante.nombre).border = thin_border
        ws2.cell(row=row, column=2, value=cal.jurado.nombre).border = thin_border
        ws2.cell(row=row, column=3, value=cal.criterio.nombre).border = thin_border
        ws2.cell(row=row, column=4, value=cal.criterio.componente).border = thin_border
        ws2.cell(row=row, column=5, value=cal.nota).border = thin_border
        row += 1
    for col in range(1, 6):
        ws2.column_dimensions[chr(64 + col)].width = 28
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='evaluacion_hilos_de_la_tierra.xlsx', as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/estudiantes')
def api_estudiantes():
    estudiantes = Estudiante.query.order_by(Estudiante.nombre).all()
    return jsonify([{'id': e.id, 'nombre': e.nombre, 'codigo': e.codigo, 'grupo': e.grupo} for e in estudiantes])

@app.route('/api/calificacion/<int:estudiante_id>/<int:jurado_id>')
def api_calificacion(estudiante_id, jurado_id):
    resultado = calcular_notas(estudiante_id, jurado_id)
    return jsonify(resultado)

# ─── Rutas de administraci├│n ───

@app.route('/admin')
@login_required
@admin_required
def admin_index():
    total_estudiantes = Estudiante.query.count()
    total_jurados = Jurado.query.count()
    total_grupos = Grupo.query.count()
    total_calificaciones = Calificacion.query.count()
    return render_template('admin/index.html', total_estudiantes=total_estudiantes,
                           total_jurados=total_jurados, total_grupos=total_grupos,
                           total_calificaciones=total_calificaciones)

# ─── CRUD Grupos ───

@app.route('/admin/grupos')
@login_required
@admin_required
def admin_grupos():
    grupos = Grupo.query.order_by(Grupo.nombre).all()
    return render_template('admin/grupos.html', grupos=grupos)

@app.route('/admin/grupos/nuevo', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_grupo_nuevo():
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        descripcion = request.form.get('descripcion', '').strip()
        horario = request.form.get('horario', '').strip()
        if not nombre:
            flash('El nombre del grupo es obligatorio', 'error')
        elif Grupo.query.filter_by(nombre=nombre).first():
            flash('Ya existe un grupo con ese nombre', 'error')
        else:
            grupo = Grupo(nombre=nombre, descripcion=descripcion, horario=horario if horario else None)
            db.session.add(grupo)
            db.session.commit()
            flash(f'Grupo "{nombre}" creado correctamente', 'success')
            return redirect(url_for('admin_grupos'))
    return render_template('admin/grupo_form.html')

@app.route('/admin/grupos/<int:grupo_id>/editar', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_grupo_editar(grupo_id):
    grupo = db.session.get(Grupo, grupo_id)
    if not grupo:
        flash('Grupo no encontrado', 'error')
        return redirect(url_for('admin_grupos'))
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        descripcion = request.form.get('descripcion', '').strip()
        horario = request.form.get('horario', '').strip()
        if not nombre:
            flash('El nombre del grupo es obligatorio', 'error')
        else:
            existente = Grupo.query.filter_by(nombre=nombre).first()
            if existente and existente.id != grupo_id:
                flash('Ya existe otro grupo con ese nombre', 'error')
            else:
                grupo.nombre = nombre
                grupo.descripcion = descripcion
                grupo.horario = horario if horario else None
                db.session.commit()
                flash(f'Grupo "{nombre}" actualizado', 'success')
                return redirect(url_for('admin_grupos'))
    return render_template('admin/grupo_form.html', grupo=grupo)

@app.route('/admin/grupos/<int:grupo_id>/eliminar', methods=['POST'])
@login_required
@admin_required
def admin_grupo_eliminar(grupo_id):
    grupo = db.session.get(Grupo, grupo_id)
    if not grupo:
        flash('Grupo no encontrado', 'error')
    else:
        db.session.delete(grupo)
        db.session.commit()
        flash(f'Grupo "{grupo.nombre}" eliminado', 'success')
    return redirect(url_for('admin_grupos'))

# ─── CRUD Estudiantes ───

@app.route('/admin/estudiantes')
@login_required
@admin_required
def admin_estudiantes():
    grupos = Grupo.query.order_by(Grupo.nombre).all()
    estudiantes_por_grupo = {}
    for g in grupos:
        estudiantes_por_grupo[g] = Estudiante.query.filter_by(grupo_id=g.id).order_by(Estudiante.nombre).all()
    sin_grupo = Estudiante.query.filter_by(grupo_id=None).order_by(Estudiante.nombre).all()
    return render_template('admin/estudiantes.html', estudiantes_por_grupo=estudiantes_por_grupo, sin_grupo=sin_grupo)

@app.route('/admin/estudiantes/cargar', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_estudiantes_cargar():
    if request.method == 'POST':
        archivo = request.files.get('archivo')
        if not archivo or archivo.filename == '':
            flash('Selecciona un archivo Excel', 'error')
            return redirect(url_for('admin_estudiantes_cargar'))
        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            filas = list(ws.iter_rows(min_row=2, values_only=True)) if ws.max_row > 1 else []
            if not filas:
                flash('El archivo no tiene datos (la primera fila se usa como encabezado)', 'error')
                return redirect(url_for('admin_estudiantes_cargar'))
            creados = 0
            actualizados = 0
            grupos_creados = 0
            for row in filas:
                if len(row) < 3:
                    continue
                codigo = str(row[0]).strip() if row[0] is not None else ''
                nombre = str(row[1]).strip() if row[1] is not None else ''
                grupo_nombre = str(row[2]).strip() if row[2] is not None else ''
                if not codigo or not nombre:
                    continue
                grupo_obj = None
                if grupo_nombre:
                    grupo_obj = Grupo.query.filter_by(nombre=grupo_nombre).first()
                    if not grupo_obj:
                        grupo_obj = Grupo(nombre=grupo_nombre, descripcion=f'Importado de Excel')
                        db.session.add(grupo_obj)
                        db.session.flush()
                        grupos_creados += 1
                estudiante = Estudiante.query.filter_by(codigo=codigo).first()
                if estudiante:
                    estudiante.nombre = nombre
                    estudiante.grupo_id = grupo_obj.id if grupo_obj else None
                    estudiante.grupo = grupo_obj.nombre if grupo_obj else None
                    actualizados += 1
                else:
                    estudiante = Estudiante(
                        nombre=nombre, codigo=codigo,
                        grupo_id=grupo_obj.id if grupo_obj else None,
                        grupo=grupo_obj.nombre if grupo_obj else None
                    )
                    db.session.add(estudiante)
                    creados += 1
            db.session.commit()
            msg = f'Importación completada: {creados} creados, {actualizados} actualizados'
            if grupos_creados:
                msg += f', {grupos_creados} grupos nuevos'
            flash(msg, 'success')
            return redirect(url_for('admin_estudiantes'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al procesar el archivo: {str(e)}', 'error')
            return redirect(url_for('admin_estudiantes_cargar'))
    return render_template('admin/estudiantes_cargar.html')

@app.route('/admin/estudiantes/nuevo', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_estudiante_nuevo():
    grupos = Grupo.query.order_by(Grupo.nombre).all()
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        codigo = request.form.get('codigo', '').strip()
        grupo_id = request.form.get('grupo_id')
        grupo_id = int(grupo_id) if grupo_id and grupo_id.isdigit() else None
        if not nombre or not codigo:
            flash('Nombre y c├│digo son obligatorios', 'error')
        elif Estudiante.query.filter_by(codigo=codigo).first():
            flash('Ya existe un estudiante con ese c├│digo', 'error')
        else:
            grupo_obj = db.session.get(Grupo, grupo_id) if grupo_id else None
            estudiante = Estudiante(nombre=nombre, codigo=codigo, grupo_id=grupo_id, grupo=grupo_obj.nombre if grupo_obj else None)
            db.session.add(estudiante)
            db.session.commit()
            flash(f'Estudiante "{nombre}" creado correctamente', 'success')
            return redirect(url_for('admin_estudiantes'))
    return render_template('admin/estudiante_form.html', grupos=grupos)

@app.route('/admin/estudiantes/<int:estudiante_id>/editar', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_estudiante_editar(estudiante_id):
    estudiante = db.session.get(Estudiante, estudiante_id)
    if not estudiante:
        flash('Estudiante no encontrado', 'error')
        return redirect(url_for('admin_estudiantes'))
    grupos = Grupo.query.order_by(Grupo.nombre).all()
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        codigo = request.form.get('codigo', '').strip()
        grupo_id = request.form.get('grupo_id')
        grupo_id = int(grupo_id) if grupo_id and grupo_id.isdigit() else None
        if not nombre or not codigo:
            flash('Nombre y c├│digo son obligatorios', 'error')
        else:
            existente = Estudiante.query.filter_by(codigo=codigo).first()
            if existente and existente.id != estudiante_id:
                flash('Ya existe otro estudiante con ese c├│digo', 'error')
            else:
                grupo_obj = db.session.get(Grupo, grupo_id) if grupo_id else None
                estudiante.nombre = nombre
                estudiante.codigo = codigo
                estudiante.grupo_id = grupo_id
                estudiante.grupo = grupo_obj.nombre if grupo_obj else None
                db.session.commit()
                flash(f'Estudiante "{nombre}" actualizado', 'success')
                return redirect(url_for('admin_estudiantes'))
    return render_template('admin/estudiante_form.html', estudiante=estudiante, grupos=grupos)

@app.route('/admin/estudiantes/<int:estudiante_id>/eliminar', methods=['POST'])
@login_required
@admin_required
def admin_estudiante_eliminar(estudiante_id):
    estudiante = db.session.get(Estudiante, estudiante_id)
    if not estudiante:
        flash('Estudiante no encontrado', 'error')
    else:
        db.session.delete(estudiante)
        db.session.commit()
        flash(f'Estudiante "{estudiante.nombre}" eliminado', 'success')
    return redirect(url_for('admin_estudiantes'))

# ─── CRUD Jurados ───

@app.route('/admin/jurados')
@login_required
@admin_required
def admin_jurados():
    jurados = Jurado.query.order_by(Jurado.nombre).all()
    return render_template('admin/jurados.html', jurados=jurados)

@app.route('/admin/jurados/nuevo', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_jurado_nuevo():
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        if not nombre:
            flash('El nombre del jurado es obligatorio', 'error')
        else:
            jurado = Jurado(nombre=nombre, email=email if email else None)
            if password:
                jurado.password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            db.session.add(jurado)
            db.session.commit()
            flash(f'Jurado "{nombre}" creado correctamente', 'success')
            return redirect(url_for('admin_jurados'))
    return render_template('admin/jurado_form.html')

@app.route('/admin/jurados/<int:jurado_id>/editar', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_jurado_editar(jurado_id):
    jurado = db.session.get(Jurado, jurado_id)
    if not jurado:
        flash('Jurado no encontrado', 'error')
        return redirect(url_for('admin_jurados'))
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        if not nombre:
            flash('El nombre del jurado es obligatorio', 'error')
        else:
            jurado.nombre = nombre
            jurado.email = email if email else None
            if password:
                jurado.password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            db.session.commit()
            flash(f'Jurado "{nombre}" actualizado', 'success')
            return redirect(url_for('admin_jurados'))
    return render_template('admin/jurado_form.html', jurado=jurado)

@app.route('/admin/jurados/<int:jurado_id>/eliminar', methods=['POST'])
@login_required
@admin_required
def admin_jurado_eliminar(jurado_id):
    jurado = db.session.get(Jurado, jurado_id)
    if not jurado:
        flash('Jurado no encontrado', 'error')
    else:
        db.session.delete(jurado)
        db.session.commit()
        flash(f'Jurado "{jurado.nombre}" eliminado', 'success')
    return redirect(url_for('admin_jurados'))

if __name__ == '__main__':
    init_db()
    app.run(debug=True)
